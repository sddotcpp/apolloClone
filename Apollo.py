from bs4 import BeautifulSoup
from time import sleep
from lxml import html
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import re
from openpyxl import Workbook, load_workbook

#--Class Apollo.io--
class Apollo:

    #--initiator--
    def __init__(self, email, password):
        self.credentials = [email, password]
        self.options = Options()
        self.driver = webdriver.Firefox(options=self.options)
        self.EMAIL_REGEX = r"[\w\.-]+@[\w\.-]+"
    

    #--open--
    def open(self):
        c = 0
        while True and c < 10:
            try:
                c += 1
                self.driver.get('https://app.apollo.io/#/login')
                sleep(5)
                break
            except BaseException as e:
                print(e, "open Method")

        
    #--open--
    def headless(self, boolValue) -> bool:
         c = 0
         while True and c < 10:
            try:
                c += 1
                if boolValue == True:
                    self.options.add_argument('--headless')
                else:
                    pass
                break
            except BaseException as e:
                print(e, "headless Method")
    


    #--log in--
    def login(self):
        c = 0
        while True and c < 10:
            try:
                c += 1
                self.Email = self.driver.find_element(By.ID, 'o1-input')
                self.Pass =  self.driver.find_element(By.ID, 'o2-input')
                self.butt =  self.driver.find_element(By.XPATH, "//div[text() = 'Log In']")
                sleep(2)
                break
            except BaseException as e:
                print(e, "login Method 01")
            
        while True and c < 10:
            try:
                c += 1
                self.Email.send_keys("amol.agarwal@goodera.com")
                self.Pass.send_keys("DataEntry@1234")
                sleep(2)
                self.butt.click()
                sleep(5)
                break
            except BaseException as e:
                print(e, "login Method 01")
    

    #--Open search page--
    def openSearch(self):
        
        c = 0
        while True and c < 10:
            try:
                c += 1
                self.searchButt = self.driver.find_element(By.ID, 'searcher')
                self.searchButt.click()
                sleep(5)
                break
            except BaseException as e:
                print(e, "openSearch Method")

    
    #--Main search Method-- *it clears and search one time, then return the HTML *
    def search(self, Query):
        
        c = 0
        while True and c < 10:
            try:
                c += 1
                #-defin the elements-
                self.searchBar = self.driver.find_element(By.XPATH, '/html/body/div[1]/div/div[2]/div[2]/div/div[1]/div/div[2]/div/div[2]/div/div/div/div/div[1]/div/div[1]/div/div/div/div/input')
                # self.searchBar = self.driver.find_element(By.ID, 'o4-input')
                # if self.searchBar.get_attribute('placeholder') != 'Search People...':
                #     self.searchBar = self.driver.find_element(By.ID, 'o5-input')                    
                break
            except BaseException as e:
                print(e, "search Method 01")
        
        c = 0
        while True and c < 10:
            try:              
                c += 1
                #-put the query-
                self.searchBar.clear()
                self.searchBar.send_keys(Query)
                self.searchBar.send_keys(Keys.ENTER)
                sleep(5)
                break
                
            except BaseException as e:
                print(e, "search Method 02")
        self.pageSourse = self.driver.page_source

        return self.pageSourse
    

    #--get names and links--
    def getPdata(self, Html):
        self.Data = []
        spanCounter = 0
        span = ""
        c = 0
        while True and c < 10:
            try:
                c += 1
                parsed = BeautifulSoup(Html, 'html.parser')
                for table in parsed.find_all('table', class_ = ""):
                    for a in table.find_all('a'):
                        link = str(a.get('href'))
                        name = str(a.text).strip()
                        #-formating the names-
                        if len(name) > 50:
                            temp = str()
                            for char in name: 
                                if char == ' ':
                                    pass
                                elif char == '\n':
                                    temp = temp + " "
                                else:
                                    temp = temp + char
                            name = temp
                        else :
                            pass
                        
                
                        if link.startswith('#/people') or link.startswith('#/contacts'):
                            self.Data.append([name, f'https://app.apollo.io/{link}'])
                            
                    
                    for span in table.find_all('span', class_ = 'zp_1YzpM'):
                        try:
                            if spanCounter == len(self.Data):
                                break
                            self.Data[spanCounter].append(span.text)
                            spanCounter += 1 
                        except:
                            self.Data[spanCounter].append("Error getting the title")
                


                break
            except BaseException as e:
                print(e, "getPdata Method")
        
        return self.Data
    

    #--Indv Pages Data--
    def getSdata(self, PgaeUrl, useCredits = False):
        c = 0
        while True and c < 10:
            try:
                c += 1
                
                #-open the page-
                self.driver.get(PgaeUrl)
                sleep(5)

                
                #-manging the data | Emails-
                if useCredits == False:
                    HtmlText = self.driver.page_source
                    if "Access Email" in HtmlText:
                        emailsList = ["Acess Needed"]
                    else: 
                        emailsList = re.findall(self.EMAIL_REGEX, HtmlText)
                elif useCredits == True:
                    HtmlText = self.driver.page_source
                    if "Access Email" in HtmlText:
                        butt01 = self.driver.find_element(By.XPATH, "/html/body/div[1]/div/div[2]/div[2]/div/div[1]/div/div[2]/div/div/div/div/div[1]/div[1]/div/div/div/div/div[1]/div/div[2]/div/button")
                        butt01.click()
                        sleep(2)
                        
                        check02 = self.driver.find_element(By.XPATH, "/html/body/div[5]/div[2]/div/div/div[2]/form/div/div/div/div/div/div/label/div")
                        check02.click()
                        sleep(2)
                        
                        butt02 = self.driver.find_element(By.XPATH, "/html/body/div[5]/div[2]/div/div/div[3]/div")
                        butt02.click()
                        sleep(2)

                        #-after those steps, the email will be there-
                        #-the new html text-
                        HtmlText = self.driver.page_source
                        emailsList = re.findall(self.EMAIL_REGEX, HtmlText)
                    else: 
                        emailsList = re.findall(self.EMAIL_REGEX, HtmlText)    
                
                

                goToSearch = self.driver.find_element(By.XPATH, '/html/body/div[1]/div/div[2]/div[2]/div/div[1]/div/div[1]/div[1]/div[1]/div/div[1]/a')
                goToSearch.click()
                sleep(5)
                break
            except BaseException as e:
                print(e, "open Method")

        return emailsList


    #--write--
    def write(self, list_of_lists):
        self.Book = Workbook()
        self.sheet = self.Book.active
        self.sheet.title = "Data"

        for row in list_of_lists:
            try:
                self.sheet.append(row[0]) #The zero index here is to fix an extra [] issue
            except:
                self.sheet.append(["Error Happeneed in he writing method"])
        self.Book.save(r'C:\Users\Ahmed Mohamed\Desktop\Apollo.io\result.xlsx')

    #--read--
    def read(self, FilePath):
        self.Book = load_workbook(FilePath)
        self.sheet = self.Book.active

        self.queries =[]

        for row in self.sheet.values:
            self.queries.append(row[0])
        
        return self.queries


#-----------------------------Main Method--------------------------------#

    def collected(self, queriesList, searchDeep = 1):
        self.Urls = []
        self.Names = []
        self.collective = []
        self.row = []

        self.open()
        # self.headless(True)
        self.login()
        self.openSearch()
        for query in queriesList:
            Data = self.getPdata(self.search(query))
            
            while True:
                try:
                    FilteredData = Data[:searchDeep]
                    break
                except IndexError:
                    searchDeep = searchDeep - 1
                    pass
                
            for ele in FilteredData: 
                self.row.append(ele)
                emails = self.getSdata(ele[1])
                if len(emails) > 0:
                    for email in emails:
                        self.row[0].append(email)
                self.collective.append(self.row)
                self.row = []
        
        self.write(self.collective)
        return self.collective
    




inst = Apollo('amol.agarwal@goodera.com', 'DataEntry@1234')

inst.collected(inst.read("test.xlsx"), 1)
